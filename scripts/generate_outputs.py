#!/usr/bin/env python3
"""
generate_outputs.py — Deterministic output generator for COCHE Paper Tracker
Reads coche_pubmed.json and produces ALL output files:
  README.md, FULL_LIST.md, COCHE_Weekly_Report.md, COCHE_Papers.xlsx, index.md

Usage: python3 scripts/generate_outputs.py [--workspace /path/to/workspace]

=== ITC Reporting Rules (2025-07 update) ===
  - Channel B = papers containing the word "InnoHK" (not ITC full name)
  - ITC reporting period: Dec 1 – Nov 30 (e.g., 2024-12-01 to 2025-11-30 = ITC 2025)
  - Sorting: by pub_year DESC, pub_month DESC, pub_day DESC
  - "Recent 30 days" = papers where date >= (today - 30 days)
  - Excel: full author names, corresponding author(s) marked with *

Rules (DO NOT MODIFY):
  - All dates are extracted from JSON fields as-is
  - This script contains ZERO LLM-generated content — only deterministic logic
"""

import json
import os
import re
import sys
from collections import Counter
from datetime import datetime, timedelta

WORKSPACE = os.environ.get('COCHE_WORKSPACE', '/home/ubuntu/.openclaw/workspace')
args = sys.argv[1:]
for i, arg in enumerate(args):
    if arg == '--workspace' and i + 1 < len(args):
        WORKSPACE = args[i + 1]

PUBMED_FILE = os.path.join(WORKSPACE, 'coche_pubmed.json')
PREV_FILE = os.path.join(WORKSPACE, 'coche_pubmed_previous.json')
README_FILE = os.path.join(WORKSPACE, 'README.md')
FULL_LIST_FILE = os.path.join(WORKSPACE, 'FULL_LIST.md')
REPORT_FILE = os.path.join(WORKSPACE, 'COCHE_Weekly_Report.md')
EXCEL_FILE = os.path.join(WORKSPACE, 'COCHE_Papers.xlsx')
INDEX_FILE = os.path.join(WORKSPACE, 'index.md')

MONTH_MAP = {'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06',
             'Jul':'07','Aug':'08','Sep':'09','Oct':'10','Nov':'11','Dec':'12'}

month_order = {'Jan':0,'Feb':1,'Mar':2,'Apr':3,'May':4,'Jun':5,
               'Jul':6,'Aug':7,'Sep':8,'Oct':9,'Nov':10,'Dec':11}

def load_papers():
    with open(PUBMED_FILE, 'r') as f:
        papers = json.load(f)
    papers.sort(key=lambda p: (
        -int(p.get('pub_year','0') or '0'),
        -month_order.get((p.get('pub_month','Jan') or 'Jan')[:3], 0),
        -int((p.get('pub_day','01') or '01'))
    ))
    return papers

def format_date(p):
    y = p.get('pub_year','') or ''
    m = (p.get('pub_month','') or 'Jan')[:3]
    d = (p.get('pub_day','') or '01').zfill(2)
    mm = MONTH_MAP.get(m, '01')
    return f'{y}-{mm}-{d}'

def format_link(p):
    doi = p.get('doi','')
    if doi: return f'https://doi.org/{doi}'
    return f'https://pubmed.ncbi.nlm.nih.gov/{p["pmid"]}'

def format_authors(p, max_n=3, suffix=' et al.'):
    authors = p.get('coche_authors', [])
    if not authors: return 'N/A'
    shown = ','.join(authors[:max_n])
    if len(authors) > max_n: return f'{shown}{suffix}'
    return shown

def format_all_authors_with_corresponding(p):
    """
    Return a comma-separated string of ALL authors.
    Corresponding author(s) are marked with *.
    
    Detection priority:
      1. author_list with is_corresponding=True (from PubMed XML or EZproxy scan)
      2. Fallback: if no corresponding author found, mark the last author(s) 
         (common convention: corresponding author is often the last author)
      3. Old-style flat list of strings
    """
    author_data = p.get('author_list', [])
    if not author_data:
        # Fallback: old-style flat list of strings
        flat = p.get('authors', [])
        if flat and isinstance(flat[0], str):
            return ', '.join(flat)
        elif flat and isinstance(flat[0], dict):
            return ', '.join(a.get('name', '') for a in flat)
        return ''

    # Check if any corresponding author is already marked
    has_marked = any(a.get('is_corresponding', False) for a in author_data)
    
    parts = []
    for i, a in enumerate(author_data):
        name = a.get('name', '').strip()
        if not name:
            continue
        is_corr = a.get('is_corresponding', False)
        # Fallback: if no corresponding author is marked in the entire list,
        # mark the last 1-2 author(s) as corresponding (common for papers with
        # joint corresponding authors). The exact count depends on author list length:
        # 1-5 authors: mark last 1; 6+: mark last 2 (when no explicit markers exist)
        if not has_marked:
            fallback_count = 1 if len(author_data) <= 5 else 2
            corr_indices = set(range(len(author_data) - fallback_count, len(author_data)))
        else:
            corr_indices = set()
        if i in corr_indices:
            is_corr = True
        if is_corr:
            name = f"{name}*"
        parts.append(name)
    return ', '.join(parts)

def get_itc_reporting_year(p):
    """
    Compute the ITC reporting year for a paper.
    ITC cycle: Dec 1 to Nov 30.
    E.g., paper published 2025-12-15 → ITC 2026
           paper published 2025-11-15 → ITC 2025
           paper published 2025-06-01 → ITC 2025
    Returns (reporting_year, sort_key) where sort_key is (year, month, day)
    """
    y = int(p.get('pub_year', '0') or '0')
    m = (p.get('pub_month', 'Jan') or 'Jan')[:3]
    d = int((p.get('pub_day', '01') or '01'))
    month_num = month_order.get(m, 0) + 1  # 1-based

    # If month is Dec, it belongs to NEXT year's ITC cycle
    if month_num == 12:
        itc_year = y + 1
    else:
        itc_year = y
    return itc_year, (y, month_order.get(m, 0), d)

def is_channel_b(p):
    """
    Channel B: paper must contain the word 'InnoHK' in its acknowledgement.
    Previously this also matched ITC full name — now restricted to InnoHK only.
    """
    return 'innohk_acknowledgement' in p.get('source', [])

def is_dual(p):
    """Dual = has both COCHE affiliation AND InnoHK acknowledgement."""
    return ('affiliation' in p.get('source', [])) and is_channel_b(p)

def is_recent(p, threshold_date):
    return format_date(p) >= threshold_date

def get_innohk_snippet(p):
    """Get the InnoHK acknowledgement snippet from the paper."""
    snippet = p.get('innohk_snippet', '')
    if not snippet:
        return ''
    # Truncate and clean
    if len(snippet) > 300:
        snippet = snippet[:300] + '...'
    return snippet

def generate_readme(papers, recent, now):
    innohk_total = sum(1 for p in papers if is_channel_b(p))
    innohk_dual = sum(1 for p in papers if is_dual(p))
    recent_inno = [p for p in recent if is_dual(p)]

    lines = []
    lines.append('# 🧠 COCHE Paper Tracker')
    lines.append('')
    lines.append('> **Hong Kong Centre for Cerebro-Cardiovascular Health Engineering**  ')
    lines.append(f'> 📊 **{len(papers)} papers** | ⭐ InnoHK: {innohk_total} ({innohk_dual} dual) | ⏰ {now.strftime("%Y-%m-%d %H:%M")} UTC+8')
    lines.append('')
    lines.append('📥 [Excel](COCHE_Papers.xlsx) | [JSON](coche_pubmed.json) | [Full Table](FULL_LIST.md) | [Report](COCHE_Weekly_Report.md)')
    lines.append('')
    lines.append('## ⭐ InnoHK Filter')
    lines.append('')
    lines.append('**⭐ InnoHK** = Paper text contains the word "InnoHK" in its acknowledgements. This is the primary ITC KPI reporting metric.')
    lines.append('')
    lines.append('- Papers with **both** COCHE affiliation AND InnoHK acknowledgement are shown prominently')
    lines.append('- Papers with InnoHK but no COCHE affiliation (InnoHK-only) are also relevant')
    lines.append('- Papers with only COCHE affiliation (no InnoHK) are collapsed by default')
    lines.append('')
    lines.append('---')
    lines.append('')

    # ⭐ Tier 1: InnoHK dual — always prominent
    if recent_inno:
        lines.append(f'## ⭐ Recent — InnoHK ({len(recent_inno)})')
        lines.append('')
        for i, p in enumerate(recent_inno, 1):
            lines.append(f'{i}. **[{p["title"]}]({format_link(p)})**')
            lines.append(f'   - 📅 {format_date(p)} | 📰 {(p.get("journal","") or "N/A")[:40]} | 👤 {format_authors(p)}')
        lines.append('')

    # 📋 Other recent papers — collapsed
    other_recent = [p for p in recent if not is_dual(p)]
    if other_recent:
        lines.append(f'<details><summary>📋 Other Recent ({len(other_recent)} papers, no InnoHK or single-channel)</summary>')
        lines.append('')
        for p in other_recent:
            tag = ' 🏷 InnoHK-only' if is_channel_b(p) else ''
            lines.append(f'- [{p["title"][:80]}]({format_link(p)}) ({format_date(p)}){tag}')
        lines.append('')
        lines.append('</details>')
        lines.append('')

    # Full list by ITC reporting year
    lines.append('## 📋 All Papers by ITC Year')
    lines.append('')
    itc_groups = {}
    for p in papers:
        itc_yr, sort_key = get_itc_reporting_year(p)
        itc_groups.setdefault(itc_yr, []).append(p)
    for itc_yr in sorted(itc_groups.keys(), reverse=True):
        yr = itc_groups[itc_yr]
        yr_inno = [p for p in yr if is_dual(p)]
        yr_single = [p for p in yr if not is_dual(p)]
        yr_recent = sum(1 for p in yr if is_recent(p, threshold_str))
        badge = ''
        if yr_inno: badge += f' ⭐{len(yr_inno)}'
        if yr_recent: badge += f' 🆕{yr_recent}'
        lines.append(f'<details>')
        lines.append(f'<summary><b>ITC {itc_yr}</b> — {len(yr)} papers{badge}</summary>')
        lines.append('')
        for i, p in enumerate(yr_inno, 1):
            lines.append(f'{i}. ⭐ [{p["title"][:80]}]({format_link(p)}) — *{(p.get("journal","") or "N/A")[:30]}* ({format_date(p)})')
        if yr_single:
            yr_inno_only = [p for p in yr_single if is_channel_b(p)]
            yr_affil_only = [p for p in yr_single if not is_channel_b(p) and 'affiliation' in p.get('source',[])]
            yr_other = [p for p in yr_single if p not in yr_inno_only and p not in yr_affil_only]
            tail = ''
            if yr_affil_only: tail += f' +{len(yr_affil_only)} COCHE affiliation'
            if yr_inno_only: tail += f' +{len(yr_inno_only)} InnoHK-only'
            if yr_other: tail += f' +{len(yr_other)} other'
            lines.append(f'<details><summary>More papers{tail} (non-InnoHK)</summary>')
            for p in yr_single:
                tag = ' 🏷 InnoHK-only' if is_channel_b(p) else ''
                lines.append(f'- [{p["title"][:80]}]({format_link(p)}) ({format_date(p)}){tag}')
            lines.append('</details>')
        lines.append('')
        lines.append('</details>')
        lines.append('')
    lines.append('---')
    lines.append('> **⭐ Star = InnoHK**: paper contains the word "InnoHK" in its acknowledgements (ITC KPI metric).  ')
    lines.append('> Papers grouped by ITC reporting year (Dec 1 – Nov 30). See [FULL_LIST.md](FULL_LIST.md) for complete table.')
    lines.append('')
    lines.append('*Auto-updated weekly via PubMed API — all dates are first-published*')
    lines.append('')
    return '\n'.join(lines)

def generate_full_list(papers, now):
    innohk_total = sum(1 for p in papers if is_channel_b(p))
    innohk_dual = sum(1 for p in papers if is_dual(p))
    lines = []
    lines.append('# COCHE All Papers')
    lines.append('')
    # Group by ITC reporting year
    itc_groups = {}
    for p in papers:
        itc_year = get_itc_reporting_year(p)[0]
        itc_groups.setdefault(itc_year, []).append(p)
    
    lines.append(f'> 📊 **{len(papers)}** papers | ⭐ InnoHK: {innohk_total} ({innohk_dual} dual) | ⏰ {now.strftime("%Y-%m-%d %H:%M")} UTC+8 | [Home](README.md)')
    lines.append('')
    lines.append('> **⭐ InnoHK** = paper contains "InnoHK" in acknowledgement (ITC KPI metric). Papers are grouped by ITC reporting year (Dec 1 – Nov 30).')
    lines.append('')
    lines.append('---')
    lines.append('')
    for itc_year in sorted(itc_groups.keys(), reverse=True):
        yr_papers = itc_groups[itc_year]
        yr_inno = [p for p in yr_papers if is_channel_b(p)]
        yr_dual = [p for p in yr_papers if is_dual(p)]
        lines.append(f'## ITC {itc_year} ({len(yr_papers)} papers, {len(yr_inno)} InnoHK)')
        lines.append('')
        lines.append('| # | Title | Date | Journal | COCHE Authors | Status | PMID |')
        lines.append('|---|---|---|---|---|---|---|')
        # Sort within ITC year by pub date desc
        yr_papers_sorted = sorted(yr_papers, key=lambda p: (
            -int(p.get('pub_year','0') or '0'),
            -month_order.get((p.get('pub_month','Jan') or 'Jan')[:3], 0),
            -int((p.get('pub_day','01') or '01'))
        ))
        for i, p in enumerate(yr_papers_sorted, 1):
            title_escaped = p['title'].replace('|', '\\|')
            if is_dual(p):
                status = '⭐ InnoHK'
                prefix = f'{i}'
            elif is_channel_b(p):
                status = '🏷 InnoHK-only'
                prefix = '-'
            else:
                status = 'COCHE affiliation'
                prefix = '-'
            lines.append(f'| {prefix} | [{title_escaped[:100]}]({format_link(p)}) | {format_date(p)} | {(p.get("journal","") or "N/A")[:25]} | {format_authors(p, 99, "")} | {status} | {p["pmid"]} |')
        lines.append('')
    lines.append('---')
    lines.append('> **⭐ InnoHK** = paper has "InnoHK" acknowledgement (ITC KPI metric) | Papers grouped by ITC year (Dec–Nov) | * = corresponding author')
    lines.append(f'*Generated {now.strftime("%Y-%m-%d %H:%M:%S")}*')
    lines.append('')
    return '\n'.join(lines)

def generate_report(papers, recent, now):
    innohk_total = sum(1 for p in papers if is_channel_b(p))
    innohk_dual = sum(1 for p in papers if is_dual(p))
    innohk_only = sum(1 for p in papers if is_channel_b(p) and ('affiliation' not in p.get('source', [])))
    recent_inno = [p for p in recent if is_dual(p)]
    lines = []
    lines.append(f'# COCHE Weekly Report')
    lines.append('')
    lines.append(f'**{now.strftime("%Y-%m-%d %H:%M")}** UTC+8')
    lines.append('')
    lines.append('## Summary')
    lines.append(f'- Total: {len(papers)} | ⭐ InnoHK: {innohk_total} ({innohk_dual} + COCHE affiliation, {innohk_only} InnoHK-only)')
    lines.append(f'- Past 30d: {len(recent)} (InnoHK: {sum(1 for p in recent if is_channel_b(p))})')
    lines.append('')
    lines.append('> **⭐ InnoHK** = paper contains "InnoHK" in acknowledgements. ITC KPI metric.')
    lines.append('')
    if recent_inno:
        lines.append('## ⭐ Recent — InnoHK')
        lines.append('')
        lines.append('| # | Title | Date | Journal | COCHE Authors |')
        lines.append('|---|---|---|---|---|')
        for i, p in enumerate(recent_inno, 1):
            lines.append(f'| {i} | [{p["title"][:60]}]({format_link(p)}) | {format_date(p)} | {(p.get("journal","") or "N/A")[:25]} | {format_authors(p, 2, " …")} |')
        lines.append('')
    other_recent = [p for p in recent if not is_dual(p)]
    if other_recent:
        lines.append('## 📋 Other Recent')
        lines.append('')
        for p in other_recent:
            tag = '🏷 InnoHK-only' if is_channel_b(p) else 'COCHE affiliation'
            lines.append(f'- [{p["title"][:60]}]({format_link(p)}) | {format_date(p)} | {tag}')
        lines.append('')
    # ITC year summary
    itc_groups = {}
    for p in papers:
        itc_yr = get_itc_reporting_year(p)[0]
        itc_groups.setdefault(itc_yr, []).append(p)
    lines.append('## By ITC Reporting Year')
    lines.append('')
    lines.append('| ITC Year | Total | ⭐ InnoHK (dual) | InnoHK-only | COCHE only |')
    lines.append('|---|---|---|---|---|')
    for itc_yr in sorted(itc_groups.keys(), reverse=True):
        yr = itc_groups[itc_yr]
        yr_inno_dual = sum(1 for p in yr if is_dual(p))
        yr_inno_only = sum(1 for p in yr if is_channel_b(p) and 'affiliation' not in p.get('source',[]))
        yr_affil = sum(1 for p in yr if 'affiliation' in p.get('source',[]) and not is_channel_b(p))
        lines.append(f'| ITC {itc_yr} | {len(yr)} | {yr_inno_dual} | {yr_inno_only} | {yr_affil} |')
    lines.append('')
    lines.append('---')
    lines.append(f'*Generated {now.strftime("%Y-%m-%d %H:%M:%S")} · ⭐ = "InnoHK" in acknowledgements (ITC KPI)*')
    lines.append('')
    return '\n'.join(lines)

def generate_index(papers, recent, now):
    innohk_total = sum(1 for p in papers if is_channel_b(p))
    innohk_recent = sum(1 for p in recent if is_channel_b(p))
    lines = []
    lines.append('# COCHE Paper Tracker')
    lines.append('')
    lines.append('> **Hong Kong Centre for Cerebro-Cardiovascular Health Engineering**')
    lines.append('> Weekly auto-update · ⭐ InnoHK = paper acknowledges InnoHK (ITC KPI metric)')
    lines.append('')
    lines.append(f'📊 **Total: {len(papers)}** | ⭐ InnoHK: {innohk_total} | 🆕 Past 30d: {len(recent)} ({innohk_recent} InnoHK) | ⏰ {now.strftime("%Y-%m-%d %H:%M")}')
    lines.append('')
    lines.append('---')
    lines.append('')
    recent_inno = [p for p in recent if is_channel_b(p)]
    if recent_inno:
        lines.append(f'## ⭐ Recent — InnoHK ({len(recent_inno)})')
        lines.append('')
        lines.append('| # | Title | Date | Journal | COCHE Authors |')
        lines.append('|---|---|---|---|---|')
        for i, p in enumerate(recent_inno, 1):
            title = p['title'][:70] + ('...' if len(p['title']) > 70 else '')
            lines.append(f'| {i} | [{title}]({format_link(p)}) | {format_date(p)} | {(p.get("journal","") or "N/A")[:28]} | {format_authors(p, 2, " et al.")} |')
        lines.append('')

    lines.append(f'## 📋 By ITC Reporting Year')
    lines.append('')
    # Group by ITC year
    itc_groups = {}
    for p in papers:
        itc_yr = get_itc_reporting_year(p)[0]
        itc_groups.setdefault(itc_yr, []).append(p)
    for itc_yr in sorted(itc_groups.keys(), reverse=True):
        yr_papers = itc_groups[itc_yr]
        yr_inno = sum(1 for p in yr_papers if is_channel_b(p))
        lines.append(f'- **ITC {itc_yr}**: {len(yr_papers)} papers ({yr_inno} InnoHK)')
    lines.append('')
    lines.append('---')
    lines.append('📥 [Excel](COCHE_Papers.xlsx) | 📄 [JSON](coche_pubmed.json) | 📝 [Report](COCHE_Weekly_Report.md)')
    lines.append('')
    lines.append(f'*{now.strftime("%Y-%m-%d %H:%M:%S")} · PubMed API · ⭐ = "InnoHK" in acknowledgement (ITC KPI)*')
    lines.append('')
    return '\n'.join(lines)

def generate_excel(papers, threshold_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = 'COCHE Papers'
    hdr_font = Font(bold=True, size=11, color='FFFFFF')
    hdr_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gold_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))

    # Headers: ITC Reporting Year as an explicit column
    headers = ['#','PMID','Title','DOI','First Published','Journal','All Authors','COCHE Authors','InnoHK Status','ITC Year','Link']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align; c.border = thin_border
    col_widths = [5, 10, 55, 30, 14, 28, 55, 35, 18, 10, 35]
    for i,w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Sort: first by ITC reporting year DESC, then by pub date DESC
    papers_sorted = sorted(papers, key=lambda p: (
        -get_itc_reporting_year(p)[0],
        -get_itc_reporting_year(p)[1][0],  # year
        -get_itc_reporting_year(p)[1][1],   # month
        -get_itc_reporting_year(p)[1][2]    # day
    ))

    for idx, p in enumerate(papers_sorted, 1):
        date_str = format_date(p)
        link = format_link(p)
        itc_year = get_itc_reporting_year(p)[0]
        coche_auth = ', '.join(p.get('coche_authors', []))
        all_auth = format_all_authors_with_corresponding(p)

        is_both = is_channel_b(p) and ('affiliation' in p.get('source', []))
        is_innohk_only = is_channel_b(p) and ('affiliation' not in p.get('source', []))
        is_affil_only = ('affiliation' in p.get('source', [])) and not is_channel_b(p)

        if is_both:
            status_label = '⭐ InnoHK'
        elif is_innohk_only:
            status_label = '🏷 InnoHK-only'
        elif is_affil_only:
            status_label = 'COCHE affiliation'
        else:
            status_label = 'other'

        innohk_snip = get_innohk_snippet(p)
        if innohk_snip:
            status_label += f' | {innohk_snip}'

        row_data = [idx, p['pmid'], p['title'], p.get('doi',''), date_str,
                    p.get('journal',''), all_auth, coche_auth, status_label,
                    f'ITC {itc_year}', link]

        is_rec = date_str >= threshold_str
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=idx + 1, column=col, value=val)
            c.font = Font(size=10, bold=(is_both and is_rec),
                         color=('999999' if not is_both and not is_rec else None))
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = thin_border
            if is_rec and is_both:
                c.fill = gold_fill

    legend_row = len(papers_sorted) + 3
    ws.cell(row=legend_row, column=1, value='🟡').font = Font(size=14)
    both_count = sum(1 for p in papers if is_channel_b(p) and ('affiliation' in p.get('source',[])))
    innohk_total = sum(1 for p in papers if is_channel_b(p))
    recent_both = sum(1 for p in papers if is_channel_b(p) and ('affiliation' in p.get('source',[])) and format_date(p) >= threshold_str)
    ws.cell(row=legend_row, column=2,
            value=(f'⭐ InnoHK = papers with "InnoHK" acknowledgement ({innohk_total} total, {both_count} also have COCHE affiliation) | '
                   f'ITC Year = Dec 1–Nov 30 cycle | Gold highlight = InnoHK + past 30 days ({recent_both}) | '
                   f'Grey = non-InnoHK | * = corresponding author'))
    ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=11)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(EXCEL_FILE)
    return len(papers)

def main():
    papers = load_papers()
    now = datetime.now()
    threshold = (now - timedelta(days=30))
    global threshold_str
    threshold_str = threshold.strftime('%Y-%m-%d')
    recent = [p for p in papers if is_recent(p, threshold_str)]
    precise = sum(1 for p in papers if p.get('date_is_precise', True))
    innohk_count = sum(1 for p in papers if is_channel_b(p))
    innohk_dual = sum(1 for p in papers if is_dual(p))
    print(f"Papers: {len(papers)} total, {len(recent)} recent 30d")
    print(f"⭐ InnoHK: {innohk_count} ({innohk_dual} with COCHE affiliation)")
    print(f"Dates: {precise} precise, {len(papers)-precise} approximate")
    print(f"Threshold: {threshold_str}")
    print()
    print("Generating README.md...")
    with open(README_FILE, 'w') as f: f.write(generate_readme(papers, recent, now))
    print(f"  OK")
    print("Generating FULL_LIST.md...")
    with open(FULL_LIST_FILE, 'w') as f: f.write(generate_full_list(papers, now))
    print(f"  OK")
    print("Generating COCHE_Weekly_Report.md...")
    with open(REPORT_FILE, 'w') as f: f.write(generate_report(papers, recent, now))
    print(f"  OK")
    print("Generating index.md...")
    with open(INDEX_FILE, 'w') as f: f.write(generate_index(papers, recent, now))
    print(f"  OK")
    print("Generating COCHE_Papers.xlsx...")
    count = generate_excel(papers, threshold_str)
    xlsx_kb = os.path.getsize(EXCEL_FILE) / 1024
    print(f"  OK ({count} papers, {xlsx_kb:.0f} KB)")
    with open(PREV_FILE, 'w') as f: json.dump(papers, f, indent=None, ensure_ascii=False)
    print(f"  Snapshot saved")
    print()
    print(f"[{now.strftime('%Y-%m-%d %H:%M:%S')}] All outputs generated.")

if __name__ == '__main__':
    main()
